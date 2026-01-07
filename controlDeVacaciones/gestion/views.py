from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.http import JsonResponse 
# CORRECCIÓN 1: Asegurando que la importación de DiaFestivo sea correcta (singular)
from .models import Empleado, SaldoVacaciones, RegistroVacaciones, DiasFestivos, Departamento, ConfiguracionEmail, Notificacion
from .utils import enviar_email_nueva_solicitud, enviar_email_cambio_estado, probar_configuracion_email, crear_notificacion

from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from django.contrib import messages
from datetime import date, timedelta
from calendar import monthrange
import logging
# NUEVAS IMPORTACIONES REQUERIDAS para historial_global
from django.db.models import Sum, F 
from django import template
import sys 
import traceback
import calendar
import locale
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# IMPORTS PARA GENERAR PDF (ReportLab)
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader

# Sistema de archivos
import os

# Configuración Django
from django.conf import settings

logger = logging.getLogger(__name__)

# --- Clases y Funciones de Utilidad ---


# Definición de MESES_ESPANOL (Asegúrate de que esté definida en tu archivo)
MESES_ESPANOL = {
    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 
    7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
}


try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    try:
        # Fallback para sistemas que usan otro nombre (ej. Windows)
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
    except locale.Error:
        # Fallback si no se puede configurar
        pass 



def generar_calendario_anual_intermensual(anio):
    """
    Genera la estructura del calendario anual (Ene a Dic) con semanas agrupadas
    de 7 días consecutivos, permitiendo que las semanas crucen meses.
    
    La estructura final agrupa las semanas por el mes en el que comienzan.
    """
    calendario_anual = {}
    
    # 1. Definir el rango del año: Desde el 1 de Enero hasta el 31 de Diciembre del año dado.
    fecha_inicio = date(anio, 1, 1)
    fecha_fin = date(anio, 12, 31) 
    
    # 2. Inicializar el puntero de fecha
    fecha_actual = fecha_inicio
    
    # Usamos un bucle while para avanzar de 7 en 7
    while fecha_actual <= fecha_fin:
        
        # 3. Crear una semana (7 días)
        dias_de_la_semana = []
        
        for i in range(7):
            dia_para_agregar = fecha_actual + timedelta(days=i)
            
            # Condición de salida: si ya hemos pasado demasiado el 31 de Diciembre.
            # (Lo mantenemos flexible para que la última semana de Diciembre se complete,
            # aunque los días caigan en el siguiente año).
            if dia_para_agregar > date(anio, 12, 31) + timedelta(days=6):
                break
                
            dias_de_la_semana.append(dia_para_agregar)

        # Si se agregaron días:
        if dias_de_la_semana:
            
            # Obtener el mes de inicio de esta semana para la cabecera.
            # Usamos %b (mes abreviado, ej: 'Jan', 'Dec') y lo pasamos a mayúsculas
            mes_inicio_semana = dias_de_la_semana[0].strftime('%b').upper()
            
            # Formato de los días: Solo el número de día
            dias_solo_numero = [d.day for d in dias_de_la_semana]

            semana_data = {
                'dias': dias_solo_numero,
                'num_columnas': len(dias_solo_numero),
                # Guardamos las fechas completas para el cálculo del rango en el filtro
                'fechas_completas': dias_de_la_semana, 
            }
            
            if mes_inicio_semana not in calendario_anual:
                calendario_anual[mes_inicio_semana] = []
            
            calendario_anual[mes_inicio_semana].append(semana_data)
        
        # 4. Avanzar a la siguiente semana (7 días)
        fecha_actual = fecha_actual + timedelta(days=7)
        
        # En caso de que se haya agregado una semana que empieza en Diciembre y termina en Enero del prox año,
        # detenemos el bucle después de procesarla.
        if fecha_actual.year > anio and fecha_actual.month == 1:
            break

    return calendario_anual


from django.shortcuts import render
from datetime import date, timedelta
from calendar import monthrange
# Asumimos que tienes modelos para Empleado
# from .models import Empleado 

def generar_calendario_anual(anio):
    """
    Genera una estructura de calendario anual, dividiendo el año en semanas completas (Lunes a Domingo).
    Si una semana cae en dos meses, se asigna al mes donde cae la mayoría de los días (o el inicio).
    
    Estructura de retorno:
    {
        'ENERO': [
            {'rango': '1/Ene al 5/Ene', 'dias': [date(2025, 1, 1), ...]},
            ...
        ],
        ...
    }
    """
    calendario_anual = {}
    
    # 1. Encontrar el primer día de planificación
    # Buscamos el Lunes más cercano al 1 de Enero del año.
    fecha_inicio = date(anio, 1, 1)
    # Lunes es 0 en Python. Restamos los días necesarios.
    # Ejemplo: Si el 1 de Enero es Miércoles (2), restamos 2 días para ir al Lunes (0).
    dias_para_lunes = fecha_inicio.weekday() 
    
    # Si el 1 de enero no es lunes (0), retrocedemos para empezar el primer lunes del ciclo.
    if dias_para_lunes != 0:
         fecha_inicio -= timedelta(days=dias_para_lunes)
         
    # 2. Iterar a lo largo del año en bloques de 7 días (semanas)
    fecha_actual = fecha_inicio
    
    # Iteramos hasta que pasemos el final del año (31 de Diciembre)
    while fecha_actual.year < anio + 1:
        
        dias_semana = []
        
        # Generar los 7 días (Lunes a Domingo)
        for i in range(7):
            dias_semana.append(fecha_actual + timedelta(days=i))
            
        # El rango de días para mostrar en el encabezado
        dia_inicio = dias_semana[0]
        dia_fin = dias_semana[-1]
        
        # El mes al que pertenece esta semana (usamos el mes del día de inicio)
        nombre_mes = dia_inicio.strftime('%B').upper()
        
        # Formato del rango de días
        rango_display = f"{dia_inicio.day} al {dia_fin.day}"
        
        # Si la semana cruza meses, incluimos el mes para mayor claridad (ej: 29/Dic al 4/Ene)
        if dia_inicio.month != dia_fin.month:
            rango_display = f"{dia_inicio.day}/{dia_inicio.month} al {dia_fin.day}/{dia_fin.month}"

        # 3. Almacenar en el calendario anual
        # Inicializar la lista si es la primera vez que vemos este mes
        if nombre_mes not in calendario_anual:
            calendario_anual[nombre_mes] = []
            
        # Añadir la semana a la lista. Usamos las claves: 'rango' y 'dias'
        calendario_anual[nombre_mes].append({
            'rango': rango_display, 
            'dias': dias_semana,
        })
        
        # Mover a la siguiente semana
        fecha_actual += timedelta(days=7)

    return calendario_anual

# NOTA: La función calendario_global se encuentra más adelante en este archivo (línea ~1277)
# con la implementación completa que incluye datos reales de la base de datos.


class CustomLoginView(LoginView):
    def get_success_url(self):
        default_url = super().get_success_url()
        user = self.request.user

        if user.is_superuser:
            try:
                # CORRECCIÓN 2: El campo en Empleado es 'user', no 'usuario'
                Empleado.objects.get(user=user)
                return '/' 
            except Empleado.DoesNotExist:
                pass
        return default_url

# CORRECCIÓN ESTRUCTURAL: Se elimina la redefinición al final del archivo.
# Se usa esta versión robusta para el decorador @user_passes_test.
def is_manager(user):
    """Verifica si el usuario está autenticado y tiene un perfil de manager."""
    # Asegurarse de que el usuario tenga un perfil de empleado
    # NOTA: Si el error es de lógica de autorización, revisa qué lógica estás usando
    # (es_manager en el perfil vs. is_staff/grupos). Mantengo la versión robusta.
    return user.is_authenticated and hasattr(user, 'empleado') and user.empleado.es_manager

def calcular_dias_habiles(fecha_inicio, fecha_fin):
    """
    Calcula los días hábiles (laborables) entre dos fechas, 
    excluyendo sábados, domingos y días festivos registrados.
    """
    if fecha_inicio > fecha_fin:
        return 0
        
    delta = fecha_fin - fecha_inicio
    dias_habiles = 0
    
    # Obtener todos los días festivos en el rango para una consulta eficiente
    festivos = DiasFestivos.objects.filter(fecha__range=[fecha_inicio, fecha_fin]).values_list('fecha', flat=True)
    
    current_date = fecha_inicio
    while current_date <= fecha_fin:
        # 0 = Lunes, 6 = Domingo
        weekday = current_date.weekday()
        
        # 1. Excluir Sábados (5) y Domingos (6)
        if weekday < 5: 
            # 2. Excluir Días Festivos
            if current_date not in festivos:
                dias_habiles += 1
                
        current_date += timedelta(days=1)
        
    return dias_habiles

# --- Vistas Principales ---

@login_required
def dashboard(request):
    """
    Dashboard mejorado:
    - Para MANAGERS: Muestra KPIs del equipo (total empleados, solicitudes pendientes, etc.)
    - Para EMPLEADOS: Muestra datos personales (saldo, antigüedad, etc.)
    """
    try:
        # CORRECCIÓN 3: Se usa 'user' para filtrar el Empleado
        empleado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        if request.user.is_superuser:
            messages.error(request, "Tu cuenta de Superusuario no está vinculada a un registro de Empleado. Por favor, crea uno en el panel de Administración.")
            # Si es superusuario y no tiene perfil, lo redirige al admin
            return redirect('admin:index') 
        else:
            messages.error(request, "Error crítico: Usuario autenticado sin perfil de Empleado. Contacte a RR.HH.")
            # Si no es superusuario y no tiene perfil, lo redirige al login
            return redirect('/login/') 
            
    current_year = timezone.now().year
    context = {'empleado': empleado, 'current_year': current_year}
    
    if empleado.es_manager:
        # ============================================
        # DASHBOARD DE MANAGER - KPIs del Equipo (Filtrado por jerarquía)
        # ============================================
        from django.db.models import Q
        
        # Si es Administrador (Superuser), ve todo. Si no, solo su equipo directo + los sin manager asignado (opcional)
        # Para cumplir estrictamente "referido solo a él", filtraremos por su equipo.
        if request.user.is_superuser:
            filtro_equipo = Q() # Sin filtro para admins
            filtro_solicitudes = Q(estado=RegistroVacaciones.ESTADO_PENDIENTE)
        else:
            # Empleados que reportan directamente a él
            equipo_ids = Empleado.objects.filter(manager_aprobador=empleado).values_list('id', flat=True)
            filtro_equipo = Q(id__in=equipo_ids)
            filtro_solicitudes = Q(empleado__in=equipo_ids, estado=RegistroVacaciones.ESTADO_PENDIENTE)
        
        # 1. Total de empleados activos en su equipo
        total_empleados = Empleado.objects.filter(filtro_equipo).count()
        
        # 2. Solicitudes pendientes de su equipo
        solicitudes_pendientes = RegistroVacaciones.objects.filter(filtro_solicitudes).count()
        
        # 3. Total de días de vacaciones disponibles en su equipo
        saldos_equipo = SaldoVacaciones.objects.filter(filtro_equipo, ciclo=current_year)
        total_dias_equipo = sum(saldo.total_disponible() for saldo in saldos_equipo)
        
        # 4. Empleados de su equipo con vacaciones próximas (próximos 30 días)
        from datetime import timedelta
        hoy = date.today()
        fecha_limite = hoy + timedelta(days=30)
        
        query_proximas = Q(estado=RegistroVacaciones.ESTADO_APROBADA, fecha_inicio__gte=hoy, fecha_inicio__lte=fecha_limite)
        if not request.user.is_superuser:
            query_proximas &= Q(empleado__in=equipo_ids)
            
        vacaciones_proximas = RegistroVacaciones.objects.filter(query_proximas).select_related('empleado').order_by('fecha_inicio')[:5]
        
        # 5. Últimas solicitudes de su equipo
        ultimas_solicitudes = RegistroVacaciones.objects.filter(filtro_solicitudes).select_related('empleado').order_by('-fecha_inicio')[:5]
        
        # 6. Estadísticas de departamentos (solo de su equipo)
        departamentos_stats = []
        deptos_query = Departamento.objects.all()
        for depto in deptos_query:
            if request.user.is_superuser:
                count = Empleado.objects.filter(departamento=depto).count()
            else:
                count = Empleado.objects.filter(departamento=depto, manager_aprobador=empleado).count()
                
            if count > 0:
                departamentos_stats.append({
                    'nombre': depto.nombre,
                    'empleados': count
                })
        
        # 7. Estadísticas para gráficos (solo su equipo)
        if request.user.is_superuser:
            f_stats = Q(fecha_inicio__year=current_year)
        else:
            f_stats = Q(fecha_inicio__year=current_year, empleado__in=equipo_ids)

        estados_solicitudes = {
            'aprobadas': RegistroVacaciones.objects.filter(f_stats, estado=RegistroVacaciones.ESTADO_APROBADA).count(),
            'pendientes': RegistroVacaciones.objects.filter(f_stats, estado=RegistroVacaciones.ESTADO_PENDIENTE).count(),
            'rechazadas': RegistroVacaciones.objects.filter(f_stats, estado=RegistroVacaciones.ESTADO_RECHAZADA).count()
        }
        
        context.update({
            'es_manager_dashboard': True,
            'total_empleados': total_empleados,
            'solicitudes_pendientes': solicitudes_pendientes,
            'total_dias_equipo': total_dias_equipo,
            'vacaciones_proximas': vacaciones_proximas,
            'ultimas_solicitudes': ultimas_solicitudes,
            'departamentos_stats': departamentos_stats,
            'estados_solicitudes': estados_solicitudes,
        })
        
        # Datos personales del manager (sección secundaria)
        try:
            saldo_personal, created = SaldoVacaciones.objects.get_or_create(
                empleado=empleado,
                ciclo=current_year,
                defaults={'dias_iniciales': empleado.dias_base_lct(current_year)}
            )
            context['saldo_disponible_personal'] = saldo_personal.total_disponible()
        except Exception:
            context['saldo_disponible_personal'] = 0
    else:
        # ============================================
        # DASHBOARD DE EMPLEADO - Datos Personales
        # ============================================
        try:
            # Intenta obtener el saldo. Si no existe, se crea con LCT base por defecto
            saldo, created = SaldoVacaciones.objects.get_or_create(
                empleado=empleado, 
                ciclo=current_year,
                defaults={'dias_iniciales': empleado.dias_base_lct(current_year)}
            )
            context['saldo_disponible'] = saldo.total_disponible()
        except Exception:
            # En caso de error de DB o inicialización
            context['saldo_disponible'] = 0 
            
    return render(request, 'gestion/dashboard.html', context)


# --- VISTA para obtener saldo por AJAX ---
@login_required
def obtener_saldo_empleado(request):
    """Devuelve el saldo y KPIs de un empleado en formato JSON."""
    empleado_id = request.GET.get('empleado_id')
    year = request.GET.get('year')
    
    # Usar año actual si no se especifica
    if year:
        try:
            current_year = int(year)
        except ValueError:
            current_year = timezone.now().year
    else:
        current_year = timezone.now().year
    
    if not empleado_id:
        return JsonResponse({'error': 'Empleado ID no proporcionado'}, status=400)
    
    try:
        empleado_afectado = Empleado.objects.get(pk=empleado_id)
        
        # Intentar obtener el SaldoVacaciones SIN crearlo automáticamente
        try:
            saldo = SaldoVacaciones.objects.get(
                empleado=empleado_afectado,
                ciclo=current_year
            )
            
            saldo_disponible = saldo.total_disponible()
            
            # CORRECCIÓN: "Días Totales" y "Días Acumulados" mostrarán el saldo RESTANTE
            # "Días Totales" muestra SOLO la antigüedad (Base LCT)
            dias_totales = float(saldo.dias_iniciales or 0)
            
            # Usar el método del modelo que ya aplica la lógica FIFO (con negativos permitidos)
            dias_acumulados = saldo.dias_acumulados_restantes()
            
            # Para el desglose en el frontend:
            # Dado que Totales ahora es solo Base, y Acumulados es el resto neto...
            # Ajustamos la lógica de variables para que el frontend no se rompa, 
            # pero el display visual dependerá de lo que enviemos.
            dias_base_restantes = dias_totales # Referencia simple
            
        except SaldoVacaciones.DoesNotExist:
            # Si no existe saldo para este año, mostrar N/A
            saldo_disponible = None
            dias_totales = None
            dias_acumulados = 0
            dias_base_restantes = 0
        
        # Calcular días usados (aprobados)
        # CORRECCIÓN: Incluir vacaciones futuras (ej. 2026) que corresponden a este ciclo
        dias_usados = RegistroVacaciones.objects.filter(
            empleado=empleado_afectado,
            estado=RegistroVacaciones.ESTADO_APROBADA,
            fecha_inicio__year__gte=current_year  # Cambiado de = a >=
        ).aggregate(total=Sum('dias_solicitados'))['total'] or 0
        
        # Calcular días pendientes
        # CORRECCIÓN: Incluir vacaciones futuras (ej. 2026) que corresponden a este ciclo
        dias_pendientes = RegistroVacaciones.objects.filter(
            empleado=empleado_afectado,
            estado=RegistroVacaciones.ESTADO_PENDIENTE,
            fecha_inicio__year__gte=current_year # Cambiado de = a >=
        ).aggregate(total=Sum('dias_solicitados'))['total'] or 0
        
        return JsonResponse({
            'saldo_disponible': float(saldo_disponible) if saldo_disponible is not None else 'N/A',
            'dias_totales': float(dias_totales) if dias_totales is not None else '-',
            'dias_acumulados': float(dias_acumulados),
            'dias_base_restantes': float(dias_base_restantes) if saldo_disponible is not None else 0,
            'dias_usados': float(dias_usados),
            'dias_pendientes': float(dias_pendientes),
            'nombre_empleado': str(empleado_afectado),
            'estado': 'ok'
        })
        
    except Empleado.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Error al obtener saldo por AJAX: {e}\n{error_details}")
        return JsonResponse({
            'error': f'Error interno del servidor: {str(e)}',
            'details': error_details if request.user.is_superuser else None
        }, status=500)



@login_required
@user_passes_test(is_manager) 
def solicitar_vacaciones(request):
    """
    Vista para que el Manager/RRHH registre días de ausencia para otro empleado.
    Se registran como 'Pendiente' y no consumen saldo hasta que se aprueban.
    Ahora incluye: dias_totales, dias_usados y dias_pendientes para las cards KPI.
    """

    current_year = timezone.now().year

    # Período de goce en Argentina
    start_goce = date(current_year, 10, 1)
    end_goce = date(current_year + 1, 4, 30)

    # Contexto base
    context = {
        'es_manager': True,
        'hoy': date.today().isoformat(),
        'todos_empleados': Empleado.objects.all().order_by('apellido', 'nombre'),
        'empleado_seleccionado_id': request.POST.get('empleado_id', request.GET.get('empleado_id')),
        'periodo_goce_inicio': start_goce.strftime("%d/%m/%Y"),
        'periodo_goce_fin': end_goce.strftime("%d/%m/%Y"),
        'estado_a_registrar': 'Pendiente',
        'saldo_disponible': 'N/A'
    }

    empleado_id_inicial = context['empleado_seleccionado_id']

    # ============================================================
    #   CARGAR SALDO Y CALCULAR LAS 3 KPI (totales / usados / pendientes)
    # ============================================================
    if empleado_id_inicial:
        try:
            empleado_afectado = Empleado.objects.get(pk=empleado_id_inicial)

            saldo, created = SaldoVacaciones.objects.get_or_create(
                empleado=empleado_afectado,
                ciclo=current_year,
                defaults={'dias_iniciales': empleado_afectado.dias_base_lct(current_year)}
            )

            context['saldo_disponible'] = f"{saldo.total_disponible()} días"

            # -------------------------
            # KPI 1 – DÍAS TOTALES (Solo Base/Antigüedad)
            # -------------------------
            dias_totales = (saldo.dias_iniciales or 0)

            # Agregar acumulados restantes al contexto para la card
            context['dias_acumulados'] = saldo.dias_acumulados_restantes()

            # -------------------------
            # KPI 2 – DÍAS USADOS (Aprobados)
            # -------------------------
            dias_usados = RegistroVacaciones.objects.filter(
                empleado=empleado_afectado,
                estado=RegistroVacaciones.ESTADO_APROBADA,
                fecha_inicio__year=current_year
            ).aggregate(total=Sum('dias_solicitados'))['total'] or 0

            # -------------------------
            # KPI 3 – DÍAS PENDIENTES
            # -------------------------
            dias_pendientes = RegistroVacaciones.objects.filter(
                empleado=empleado_afectado,
                estado=RegistroVacaciones.ESTADO_PENDIENTE,
                fecha_inicio__year=current_year
            ).aggregate(total=Sum('dias_solicitados'))['total'] or 0

            # Agregar al contexto
            context.update({
                'dias_totales': dias_totales,
                'dias_usados': dias_usados,
                'dias_pendientes': dias_pendientes,
            })

            if created:
                messages.warning(
                    request,
                    f"Advertencia: Se inicializó automáticamente el saldo de {empleado_afectado.nombre}."
                )

        except Empleado.DoesNotExist:
            messages.error(request, f"Empleado con ID {empleado_id_inicial} no encontrado.")
        except Exception as e:
            messages.error(request, f"Error al calcular KPIs: {e}")

    # ============================================================
    #   PROCESAR FORMULARIO POST
    # ============================================================
    if request.method == 'POST':
        data = request.POST

        # 1. Empleado
        empleado_id = data.get('empleado_id')
        if not empleado_id:
            messages.error(request, "Debe seleccionar un empleado.")
            return render(request, 'gestion/solicitud.html', context)

        empleado_afectado = get_object_or_404(Empleado, pk=empleado_id)

        # 2. Fechas
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            context['empleado_seleccionado_id'] = empleado_id
            return render(request, 'gestion/solicitud.html', context)

        if fecha_inicio > fecha_fin:
            messages.error(request, "La fecha de inicio no puede ser mayor que la fecha fin.")
            context['empleado_seleccionado_id'] = empleado_id
            return render(request, 'gestion/solicitud.html', context)

        # 3. Validar período de goce - FLEXIBILIZADO: Permitir cargar vacaciones todo el año
        # if fecha_inicio < start_goce or fecha_fin > end_goce:
        #     messages.error(
        #         request,
        #         f"Fechas fuera del período de goce ({start_goce.strftime('%d/%m/%Y')} al {end_goce.strftime('%d/%m/%Y')})."
        #     )
        #     context['empleado_seleccionado_id'] = empleado_id
        #     return render(request, 'gestion/solicitud.html', context)

        # 4. Obtener saldo actualizado
        saldo, created = SaldoVacaciones.objects.get_or_create(
            empleado=empleado_afectado,
            ciclo=current_year,
            defaults={'dias_iniciales': empleado_afectado.dias_base_lct(current_year)}
        )

        saldo_disponible = saldo.total_disponible()

        dias_solicitados = (fecha_fin - fecha_inicio).days + 1

        if dias_solicitados > saldo_disponible:
            messages.error(
                request,
                f"Saldo insuficiente. Solicita {dias_solicitados} días y solo tiene {saldo_disponible} disponibles."
            )
            context['empleado_seleccionado_id'] = empleado_id
            return render(request, 'gestion/solicitud.html', context)

        # 5. Guardar solicitud
        try:
            with transaction.atomic():
                solicitud = RegistroVacaciones.objects.create(
                    empleado=empleado_afectado,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    estado=RegistroVacaciones.ESTADO_PENDIENTE,
                    manager_aprobador=None,
                    razon=data.get('razon')
                )
                
                # Enviar notificación por email
                enviar_email_nueva_solicitud(request, solicitud)

                # Notificación interna para el Administrador Global (Superusuario) y para el manager
                # (En este caso el manager ya lo sabe porque él mismo la creó, pero la dejamos para trazabilidad)
                admin_users = User.objects.filter(is_superuser=True)
                for a_user in admin_users:
                    crear_notificacion(
                        usuario=a_user,
                        titulo="Nueva Solicitud Creada (Manager) 🔔",
                        mensaje=f"Se ha registrado una solicitud para {empleado_afectado.nombre} {empleado_afectado.apellido}.",
                        url="gestion:aprobacion_manager",
                        solicitud=solicitud
                    )

            messages.success(
                request,
                f"Solicitud registrada: {dias_solicitados} días naturales. Queda en estado PENDIENTE."
            )
            return redirect('gestion:historial_global')

        except Exception as e:
            messages.error(request, f"Error al guardar solicitud: {e}")
            context['empleado_seleccionado_id'] = empleado_id
            return render(request, 'gestion/solicitud.html', context)

    # ============================================================
    #   GET → Render inicial
    # ============================================================
    return render(request, 'gestion/solicitud.html', context)



@login_required
@user_passes_test(is_manager)
def crear_empleado(request):
    departamentos = Departamento.objects.all().order_by('nombre')
    # Corregido: 'manager_aprobador' era antes 'reporta_a' en el código original, 
    # pero el modelo usa manager_aprobador
    managers_list = Empleado.objects.filter(es_manager=True).order_by('apellido')
    
    # Se usa 'user' en lugar de 'usuario_id' para la exclusión
    usuarios_con_perfil_existente = Empleado.objects.all().values_list('user__id', flat=True)
    usuarios_disponibles = User.objects.filter(is_active=True).exclude(id__in=usuarios_con_perfil_existente).order_by('username')
    
    contexto = {
        'titulo': 'Registrar Nuevo Empleado',
        'departamentos': departamentos,
        'managers_list': managers_list,
        'usuarios_disponibles': usuarios_disponibles,
    }

    if request.method == 'POST':
        data = request.POST
        errores = []
        if not data.get('usuario'):
            errores.append("Debe seleccionar un Usuario del Sistema.")
        # ... (otras validaciones) ...

        if errores:
            for error in errores:
                messages.error(request, error)
            contexto['post_data'] = data
            return render(request, 'gestion/crear_empleado.html', contexto)

        try:
            with transaction.atomic():
                user = User.objects.get(pk=data['usuario'])
                departamento_obj = Departamento.objects.get(pk=data['departamento'])
                
                manager_aprobador_obj = None
                # Asumo que el campo en el formulario se llama 'manager_aprobador' o similar
                if data.get('manager_aprobador'):
                    manager_aprobador_obj = Empleado.objects.get(pk=data['manager_aprobador'])

                # Convertir fecha_ingreso de string a date
                fecha_ingreso_obj = datetime.strptime(data['fecha_ingreso'], '%Y-%m-%d').date()
                
                empleado_nuevo = Empleado.objects.create(
                    # CORRECCIÓN 4: Se usa 'user' para el OneToOneField
                    user=user, 
                    legajo=data['legajo'],
                    dni=data['dni'],
                    nombre=data['nombre'],
                    apellido=data['apellido'],
                    departamento=departamento_obj,
                    fecha_ingreso=fecha_ingreso_obj,
                    jornada_estandar=float(data.get('jornada_estandar', 9.0)),
                    es_manager=data.get('es_manager') == 'on',
                    manager_aprobador=manager_aprobador_obj # El modelo tiene manager_aprobador
                )
                
                # Crear saldo de vacaciones
                current_year = timezone.now().year
                
                # Verificar si se configuraron días manualmente
                if data.get('dias_iniciales'):
                    # Usar días configurados manualmente
                    dias_iniciales = float(data['dias_iniciales'])
                    dias_adicionales = float(data.get('dias_adicionales', 0))
                    
                    SaldoVacaciones.objects.create(
                        empleado=empleado_nuevo,
                        ciclo=current_year,
                        dias_iniciales=dias_iniciales,
                        dias_adicionales=dias_adicionales
                    )
                    
                    mensaje_vacaciones = f"con {dias_iniciales} días iniciales"
                    if dias_adicionales > 0:
                        mensaje_vacaciones += f" y {dias_adicionales} días adicionales"
                else:
                    # Calcular automáticamente según antigüedad (LCT)
                    dias_calculados = empleado_nuevo.dias_base_lct(current_year)
                    
                    SaldoVacaciones.objects.create(
                        empleado=empleado_nuevo,
                        ciclo=current_year,
                        dias_iniciales=dias_calculados
                    )
                    
                    mensaje_vacaciones = f"con {dias_calculados} días calculados automáticamente según antigüedad"
                
            messages.success(request, f"¡Empleado {data['nombre']} {data['apellido']} registrado con éxito! Saldo de vacaciones creado {mensaje_vacaciones}.")
            return redirect('gestion:gestion_empleados')
            
        except Exception as e:
            logger.error(f"Error interno al guardar el empleado: {e}")
            messages.error(request, f"Error interno al guardar el empleado. Detalle: {e}")
            contexto['post_data'] = data
            return render(request, 'gestion/crear_empleado.html', contexto)
    
    else:
        if not usuarios_disponibles.exists() and request.user.is_superuser:
            messages.warning(request, "No hay usuarios del sistema disponibles para asignar.")
            
        return render(request, 'gestion/crear_empleado.html', contexto)


@login_required
@user_passes_test(is_manager)
def editar_empleado(request, empleado_id):
    """Vista para editar un empleado existente"""
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    departamentos = Departamento.objects.all().order_by('nombre')
    managers_list = Empleado.objects.filter(es_manager=True).exclude(pk=empleado_id).order_by('apellido')
    
    contexto = {
        'titulo': 'Editar Empleado',
        'empleado': empleado,
        'departamentos': departamentos,
        'managers_list': managers_list,
    }

    if request.method == 'POST':
        data = request.POST
        errores = []
        
        # Validaciones básicas
        if not data.get('legajo'):
            errores.append("El legajo es obligatorio.")
        if not data.get('nombre'):
            errores.append("El nombre es obligatorio.")
        if not data.get('apellido'):
            errores.append("El apellido es obligatorio.")
        if not data.get('fecha_ingreso'):
            errores.append("La fecha de ingreso es obligatoria.")

        if errores:
            for error in errores:
                messages.error(request, error)
            return render(request, 'gestion/editar_empleado.html', contexto)

        try:
            with transaction.atomic():
                # Actualizar datos del empleado
                empleado.legajo = data['legajo']
                empleado.nombre = data['nombre']
                empleado.apellido = data['apellido']
                empleado.dni = data.get('dni', '')
                empleado.fecha_ingreso = data['fecha_ingreso']
                empleado.jornada_estandar = float(data.get('jornada_estandar', 9.0))
                empleado.es_manager = data.get('es_manager') == 'on'
                
                # Actualizar departamento si se proporcionó
                if data.get('departamento_id'):
                    empleado.departamento = Departamento.objects.get(pk=data['departamento_id'])
                else:
                    empleado.departamento = None
                
                # Actualizar manager aprobador si se proporcionó
                if data.get('manager_aprobador_id'):
                    empleado.manager_aprobador = Empleado.objects.get(pk=data['manager_aprobador_id'])
                else:
                    empleado.manager_aprobador = None
                
                empleado.save()

                # --- Lógica de Cambio de Contraseña (Solo para Managers) ---
                new_password = data.get('new_password')
                confirm_password = data.get('confirm_password')

                if empleado.es_manager and new_password:
                    if new_password == confirm_password:
                        user = empleado.user
                        user.set_password(new_password)
                        user.save()
                        messages.success(request, "Contraseña actualizada correctamente.")
                    else:
                        messages.error(request, "Las contraseñas no coinciden. No se actualizó la contraseña.")
                        # Revert atomic transaction implies raising exception, but here we might just want to warn
                        # For strictly following 'save all or nothing', we should raise an error
                        raise ValueError("Las contraseñas no coinciden")

                
            messages.success(request, f"¡Empleado {empleado.nombre} {empleado.apellido} actualizado con éxito!")
            return redirect('gestion:gestion_empleados')
            
        except Exception as e:
            logger.error(f"Error al actualizar empleado: {e}")
            messages.error(request, f"Error al actualizar el empleado: {e}")
            return render(request, 'gestion/editar_empleado.html', contexto)
    
    return render(request, 'gestion/editar_empleado.html', contexto)


@login_required
@user_passes_test(is_manager)
def eliminar_empleado(request, empleado_id):
    """Vista para eliminar un empleado"""
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    
    if request.method == 'POST':
        try:
            nombre_completo = f"{empleado.nombre} {empleado.apellido}"
            empleado.delete()
            messages.success(request, f"Empleado {nombre_completo} eliminado exitosamente.")
            return redirect('gestion:gestion_empleados')
        except Exception as e:
            logger.error(f"Error al eliminar empleado: {e}")
            messages.error(request, f"Error al eliminar el empleado: {e}")
            return redirect('gestion:gestion_empleados')
    
    contexto = {
        'empleado': empleado,
    }
    return render(request, 'gestion/eliminar_empleado.html', contexto)



@login_required
@user_passes_test(is_manager)
def calendario_manager(request):
    current_year = timezone.now().year
    start_date = date(current_year, 1, 1)
    # Ejemplo de rango, podría ser más dinámico
    end_date = date(current_year, 12, 31) 
    
    # Obtener empleados y agruparlos por departamento (usando consultas reales)
    empleados = Empleado.objects.all().select_related('departamento').order_by('departamento__nombre', 'apellido')
    empleados_por_departamento = {}
    
    for emp in empleados:
        depto_nombre = emp.departamento.nombre if emp.departamento else 'Sin Departamento'
        if depto_nombre not in empleados_por_departamento:
            empleados_por_departamento[depto_nombre] = []
        
        empleados_por_departamento[depto_nombre].append({
            'nombre': f"{emp.apellido}, {emp.nombre}", 
            'id': emp.pk
        })
    
    # Crear un rango de 150 días (por ejemplo) para la visualización del calendario
    rango_dias_visualizacion = 150 
    all_days_in_range = [start_date + timedelta(days=i) for i in range(rango_dias_visualizacion)] 
    
    context = {
        'empleados_por_departamento': empleados_por_departamento,
        'all_days_in_range': all_days_in_range,
        'current_year': current_year
    }
    return render(request, 'gestion/calendario.html', context)

@login_required
def mi_historial(request):
    # Obtener solicitudes y saldo del usuario actual
    try:
        empleado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        messages.error(request, "Perfil de empleado no encontrado.")
        return redirect('dashboard') # Redirigir al dashboard para manejo de errores
    
    current_year = timezone.now().year
    
    # Se obtienen todas las solicitudes del empleado, ordenadas por fecha de inicio
    solicitudes = RegistroVacaciones.objects.filter(empleado=empleado).order_by('-fecha_inicio')
    
    # Se obtiene el saldo asegurando el cálculo LCT si no existe
    try:
        saldo, created = SaldoVacaciones.objects.get_or_create(
            empleado=empleado, 
            ciclo=current_year,
            defaults={'dias_iniciales': empleado.dias_base_lct(current_year)}
        )
        saldo_disponible = saldo.total_disponible()
    except Exception as e:
        logger.error(f"Error al calcular saldo para empleado {empleado.user.username}: {e}")
        messages.error(request, "Error al calcular días disponibles.")
        saldo_disponible = 'ERROR'
    
    
    context = {
        'empleado': empleado,
        'solicitudes': solicitudes,
        'saldo': saldo,
        'dias_consumidos': saldo.dias_consumidos_total(),
        'anio_ciclo': current_year
    }
    return render(request, 'gestion/mi_historial.html', context)

@login_required
def mi_perfil(request):
    # Permite al usuario actualizar datos de contacto/contraseña
    # En un proyecto real, se usaría un formulario de Django
    try:
        empleado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        messages.error(request, "Perfil de empleado no encontrado.")
        return redirect('dashboard')

    context = {
        'empleado': empleado,
        'usuario': request.user # Acceso directo a datos del usuario
    }
    return render(request, 'gestion/mi_perfil.html', context)

@login_required
@user_passes_test(is_manager)
def historial_global(request):
    # --- 1. Lógica de Filtros ---
    hoy = datetime.now().date()
    
    empleado_id = request.GET.get('empleado')
    estado = request.GET.get('estado')

    # Si es Administrador (Superuser), ve todo. Si no, solo su equipo directo.
    if request.user.is_superuser:
        solicitudes_qs = RegistroVacaciones.objects.all()
        empleados_disponibles = Empleado.objects.order_by('apellido', 'nombre')
        resumen_base_qs = RegistroVacaciones.objects.all()
    else:
        equipo = Empleado.objects.filter(manager_aprobador=request.user.empleado)
        solicitudes_qs = RegistroVacaciones.objects.filter(empleado__in=equipo)
        empleados_disponibles = equipo.order_by('apellido', 'nombre')
        resumen_base_qs = RegistroVacaciones.objects.filter(empleado__in=equipo)

    # Aplicar filtros opcionales
    if empleado_id and empleado_id != 'Todos':
        solicitudes_qs = solicitudes_qs.filter(empleado__id=empleado_id)

    if estado and estado != 'Todos':
        solicitudes_qs = solicitudes_qs.filter(estado=estado)
        
    solicitudes_qs = solicitudes_qs.order_by('-fecha_solicitud')
    
    # --- 2. Datos para Filtros y Resumen ---
    
    # Cálculo del resumen (Días Aprobados)
    resumen_aprobado_data = resumen_base_qs.filter(
        estado=RegistroVacaciones.ESTADO_APROBADA, # Uso de la constante
        # Filtrar por solicitudes cuya fecha de inicio sea en el año actual
        fecha_inicio__year=datetime.now().year 
    ).values(
        'empleado__id', 
        'empleado__nombre', 
        'empleado__apellido',
        'empleado__legajo'  # <--- CAMBIO: Agregamos el campo legajo
    ).annotate(
        dias_aprobados=Sum('dias_solicitados')
    )
    
    context = {
        'solicitudes': solicitudes_qs,
        'empleados_disponibles': empleados_disponibles,
        # Necesitas definir esta tupla o lista en tu modelo RegistroVacaciones
        # EJEMPLO: ESTADOS = [('Pendiente', 'Pendiente'), ('Aprobada', 'Aprobada'), ...]
        'estados_disponibles': RegistroVacaciones.ESTADOS, 
        'empleado_id_seleccionado': empleado_id,
        'estado_seleccionado': estado,
        'resumen_aprobado': resumen_aprobado_data,
        'contexto': {'current_year': datetime.now().year},
    }
    
    # Renderizamos la plantilla
    return render(request, 'gestion/reportes.html', context)

@login_required
@user_passes_test(is_manager)
def gestion_empleados(request):
    try:
        # Si es Administrador (Superuser), ve todos. Si no, solo su equipo.
        if request.user.is_superuser:
            empleados = Empleado.objects.all().select_related('departamento', 'manager_aprobador__user').order_by('apellido', 'nombre') 
        else:
            empleados = Empleado.objects.filter(manager_aprobador=request.user.empleado).select_related('departamento', 'manager_aprobador__user').order_by('apellido', 'nombre') 
    except Exception as e:
        messages.error(request, f"Error al cargar empleados: {e}")
        empleados = [] 

    # Datos necesarios para el formulario de nuevo empleado
    departamentos = Departamento.objects.all().order_by('nombre')
    managers_list = Empleado.objects.filter(es_manager=True).order_by('apellido')
    
    # Usuarios disponibles (sin perfil de empleado asignado)
    usuarios_con_perfil_existente = Empleado.objects.all().values_list('user__id', flat=True)
    usuarios_disponibles = User.objects.filter(is_active=True).exclude(id__in=usuarios_con_perfil_existente).order_by('username')

    contexto = {
        'empleados': empleados,
        'departamentos': departamentos,
        'managers_list': managers_list,
        'usuarios_disponibles': usuarios_disponibles,
    }
    return render(request, 'gestion/gestion_empleados.html', contexto)

@login_required
@user_passes_test(is_manager)
def gestion_saldos(request):
    # Obtener todos los saldos o datos para la gestión
    current_year = timezone.now().year
    saldos = SaldoVacaciones.objects.filter(ciclo=current_year).select_related('empleado__user', 'empleado__departamento').order_by('empleado__apellido')
    
    contexto = {
        'saldos': saldos,
        'anio_ciclo': current_year
    }
    return render(request, 'gestion/saldos.html', contexto)

@login_required
@user_passes_test(is_manager)
def eliminar_festivo(request, festivo_id):
    if request.method == 'POST':
        festivo = get_object_or_404(DiasFestivos, id=festivo_id)
        festivo.delete()
        messages.success(request, 'Día festivo eliminado correctamente.')
    return redirect('gestion:gestion_festivos')

def _fetch_and_save_holidays(request, year):
    import json
    import urllib.request
    from datetime import datetime
    
    url = f"https://date.nager.at/api/v3/PublicHolidays/{year}/AR"
    count = 0
    
    try:
        # Configurar request con User-Agent
        req = urllib.request.Request(
            url, 
            data=None, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
        )
        
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
            
        for item in data:
            fecha_str = item.get('date') # YYYY-MM-DD
            nombre = item.get('localName') or item.get('name')
            fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()

            # Crear si no existe
            if not DiasFestivos.objects.filter(fecha=fecha_obj).exists():
                DiasFestivos.objects.create(fecha=fecha_obj, descripcion=nombre)
                count += 1
                
        if count > 0:
            messages.success(request, f'Se descargaron {count} días festivos para el año {year}.')
            
    except Exception as e:
        messages.warning(request, f'No se pudieron descargar los feriados de {year} automáticamente: {str(e)}')
        
    return count

@login_required
@user_passes_test(is_manager)
def gestion_festivos(request):
    if request.method == 'POST':
        fecha = request.POST.get('fecha')
        descripcion = request.POST.get('descripcion')
        
        if fecha and descripcion:
            try:
                if DiasFestivos.objects.filter(fecha=fecha).exists():
                    messages.error(request, 'Ya existe un día festivo en esta fecha.')
                else:
                    DiasFestivos.objects.create(fecha=fecha, descripcion=descripcion)
                    messages.success(request, 'Día festivo agregado correctamente.')
            except Exception as e:
                messages.error(request, f'Error al guardar: {str(e)}')
        return redirect('gestion:gestion_festivos')

    # Lógica de Filtros y Auto-Import
    today = date.today()
    anio_param = request.GET.get('anio')
    
    # Años para botones
    real_current_year = today.year
    prev_year = real_current_year - 1
    next_year = real_current_year + 1
    
    # Determinar año seleccionado (por defecto el actual)
    selected_year = None
    if anio_param == 'todo':
        selected_year = 'todo'
    elif anio_param:
        try:
            selected_year = int(anio_param)
        except ValueError:
            selected_year = real_current_year
    else:
        selected_year = real_current_year

    # Auto-importación si es un año específico y está vacío
    if isinstance(selected_year, int):
        exists = DiasFestivos.objects.filter(fecha__year=selected_year).exists()
        if not exists:
            # Intentar descargar
            _fetch_and_save_holidays(request, selected_year)

    # Query Final
    festivos = DiasFestivos.objects.all().order_by('-fecha')
    
    if isinstance(selected_year, int):
        festivos = festivos.filter(fecha__year=selected_year)

    contexto = {
        'festivos': festivos,
        'selected_year': selected_year,
        'prev_year': prev_year,
        'current_year': real_current_year,
        'next_year': next_year,
    }
    return render(request, 'gestion/festivos.html', contexto)

@login_required
@user_passes_test(is_manager)
def configurar_email(request):
    """
    Vista para configurar el servidor SMTP y los destinatarios de las notificaciones.
    Solo accesible por managers.
    """
    # Intentamos obtener la configuración con ID=1 (Singleton pattern)
    config, created = ConfiguracionEmail.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        try:
            config.email_host = request.POST.get('email_host', 'smtp.gmail.com')
            config.email_port = int(request.POST.get('email_port', 587))
            config.email_use_tls = request.POST.get('email_use_tls') == 'on'
            config.email_use_ssl = request.POST.get('email_use_ssl') == 'on'
            config.email_host_user = request.POST.get('email_host_user', '')
            config.email_host_password = request.POST.get('email_host_password', '')
            config.emails_notificacion = request.POST.get('emails_notificacion', '')
            config.activo = request.POST.get('activo') == 'on'
            config.save()
            
            # Si se presionó el botón de Probar
            if 'test' in request.POST:
                exito, mensaje = probar_configuracion_email(request, config)
                if exito:
                    messages.success(request, f"✅ Configuración Guardada: {mensaje}")
                else:
                    messages.error(request, f"❌ Error en la Prueba: {mensaje}")
            else:
                messages.success(request, "¡Configuración de email guardada con éxito!")

        except Exception as e:
            messages.error(request, f"Error al guardar la configuración: {e}")
            
        return redirect('gestion:configurar_email')

    contexto = {
        'config': config,
        'titulo_pagina': 'Configuración de Correos'
    }
    return render(request, 'gestion/configuracion_email.html', contexto)

@login_required
@user_passes_test(is_manager)
def probar_email(request):
    """
    Vista rápida para disparar un email de prueba.
    """
    config = ConfiguracionEmail.objects.filter(id=1).first()
    if not config:
        messages.error(request, "No hay configuración cargada.")
        return redirect('gestion:configurar_email')
        
    exito, mensaje = probar_configuracion_email(request, config)
    if exito:
        messages.success(request, mensaje)
    else:
        messages.error(request, f"Error de Prueba: {mensaje}")
        
    return redirect('gestion:configurar_email')





@login_required
# @user_passes_test(is_manager, login_url='/gestion/no_autorizado/')
def dias_disponibles_view(request, empleado_id=None):
    print("--- INICIANDO dias_disponibles_view ---") 
    CICLO_ACTUAL = date.today().year

    empleado_a_ver = None
    saldo = None              
    saldo_total = 0.0
    dias_tomados = 0.0          
    dias_iniciales_saldo = 0.0  
    dias_adicionales = 0.0      
    dias_otorgados = 0.0        

    # 🔹 LISTADO DE EMPLEADOS PARA EL SELECT
    empleados = Empleado.objects.all().order_by("apellido", "nombre")
    empleado_id = request.GET.get("empleado_id")

    if empleado_id:
        empleado_a_ver = get_object_or_404(Empleado, id=empleado_id)
    else:
        try:
            empleado_a_ver = Empleado.objects.get(user=request.user)
        except Empleado.DoesNotExist:
            return render(request, 'gestion/error.html', {'mensaje': 'Usuario no asociado a un empleado.'})

    try:
        saldo = SaldoVacaciones.objects.get(
            empleado=empleado_a_ver, 
            ciclo=CICLO_ACTUAL
        )

        dias_iniciales_saldo = saldo.dias_iniciales or 0.0 
        dias_adicionales = saldo.dias_iniciales or 0.0
        dias_otorgados = dias_iniciales_saldo + dias_adicionales

        
        dias_tomados_agregado = RegistroVacaciones.objects.filter(
            empleado=empleado_a_ver,
            fecha_inicio__year=CICLO_ACTUAL,
            estado=RegistroVacaciones.ESTADO_APROBADA
        ).aggregate(
            total_tomados=Sum('dias_solicitados')
        )['total_tomados']
        

        dias_tomados = dias_tomados_agregado or 0.0
        
        saldo_total = dias_otorgados - dias_tomados
        
    except SaldoVacaciones.DoesNotExist:
        dias_tomados_agregado = RegistroVacaciones.objects.filter(
            empleado=empleado_a_ver,
            fecha_inicio__year=CICLO_ACTUAL, 
            estado=RegistroVacaciones.APROBADO 
        ).aggregate(
            total_tomados=Sum('dias_utilizados')
        )['total_tomados']
        
        dias_tomados = dias_tomados_agregado or 0.0
        saldo_total = dias_otorgados - dias_tomados 
        
    context = {
        'empleado': empleado_a_ver,
        'saldo': saldo,
        'ciclo_actual': CICLO_ACTUAL, 

        'dias_iniciales_saldo': dias_iniciales_saldo,
        'dias_adicionales': dias_adicionales,
        'dias_otorgados': dias_otorgados,
        'dias_tomados': dias_tomados,
        'saldo_total': saldo_total,

        # 🔹 LISTA DE EMPLEADOS PARA MOSTRAR EN EL SELECT
        'empleados': empleados,
    }

    return render(request, 'gestion/dias_disponibles.html', context)


def total_disponible(self):
    return self.dias_iniciales + self.dias_adicionales - self.dias_consumidos_total




@login_required
@user_passes_test(is_manager)
def aprobacion_manager(request):
    """Vista para que el Manager gestione las solicitudes pendientes de su equipo."""
    try:
        empleado_manager = request.user.empleado
    except Empleado.DoesNotExist:
        # Si es un admin sin perfil de empleado, le mostramos todas las pendientes
        solicitudes_pendientes = RegistroVacaciones.objects.filter(estado=RegistroVacaciones.ESTADO_PENDIENTE)
        return render(request, 'gestion/aprobacion_manager.html', {'solicitudes_pendientes': solicitudes_pendientes})

    # Buscamos solicitudes donde el manager_aprobador sea el usuario actual
    # O solicitudes que NO tengan manager asignado (SOLO PARA ADMINS)
    from django.db.models import Q
    if request.user.is_superuser:
        filtro_gestor = Q() # Admins ven todo
    else:
        filtro_gestor = Q(manager_aprobador=empleado_manager)
        
    solicitudes_pendientes = RegistroVacaciones.objects.filter(
        filtro_gestor,
        estado=RegistroVacaciones.ESTADO_PENDIENTE
    ).order_by('fecha_solicitud')
    
    context = {
        'solicitudes_pendientes': solicitudes_pendientes
    }
    return render(request, 'gestion/aprobacion_manager.html', context)



@login_required
@user_passes_test(is_manager)
def calendario_global(request):
    """
    Vista que genera la tabla de planificación anual de vacaciones.
    Soporta vista de un año específico o todos (2024, 2025, 2026) en una sola tabla horizontal.
    """
    try:
        anio_param = request.GET.get('anio', str(date.today().year))
        
        anios_a_mostrar = []
        if anio_param == 'todos':
            anios_a_mostrar = [2024, 2025, 2026]
        else:
            try:
                anios_a_mostrar = [int(anio_param)]
            except ValueError:
                anios_a_mostrar = [date.today().year]

        # 1. GENERAR MESES Y SEMANAS CONTINUAS PARA TODOS LOS AÑOS SELECCIONADOS
        meses_globales = []
        
        for anio in anios_a_mostrar:
            datos_anio = _generar_datos_anio(anio)
            
            # Ajustar nombres de meses para incluir el año (ej: "Ene 24")
            anio_corto = str(anio)[-2:]
            for mes in datos_anio['meses_data']:
                mes['nombre'] = f"{mes['nombre']} {anio_corto}"
                meses_globales.append(mes)

        # 2. DEFINIR RANGO TOTAL DE FECHAS
        fecha_inicio_total = date(anios_a_mostrar[0], 1, 1)
        fecha_fin_total = date(anios_a_mostrar[-1], 12, 31)

        # 3. OBTENER EMPLEADOS Y VACACIONES PARA EL RANGO TOTAL
        departamentos = Departamento.objects.all().order_by('nombre')
        departamentos_data = []

        # Determinar año para el saldo (usar el actual o el primero de la lista)
        anio_saldo = date.today().year if date.today().year in anios_a_mostrar else anios_a_mostrar[0]

        for depto in departamentos:
            # Filtrar empleados por departamento Y por manager si no es superuser
            from django.db.models import Q
            if request.user.is_superuser:
                filtro_empleados = Q(departamento=depto)
            else:
                filtro_empleados = Q(departamento=depto, manager_aprobador=request.user.empleado)

            empleados_depto = Empleado.objects.filter(
                filtro_empleados
            ).select_related('manager_aprobador').order_by('apellido', 'nombre')
            
            empleados_list = []
            for emp in empleados_depto:
                # Saldo del año de referencia
                saldo, created = SaldoVacaciones.objects.get_or_create(
                    empleado=emp,
                    ciclo=anio_saldo,
                    defaults={'dias_iniciales': emp.dias_base_lct(anio_saldo)}
                )
                
                # Días acumulados: usar dias_adicionales del saldo (cargados manualmente)
                # Días base del ciclo actual (sin acumulados) - columna "Disponible"
                dias_disponibles = saldo.dias_base_ciclo()
                
                # CORRECCIÓN: Calcular consumo visualmente para incluir vacaciones puente
                consumo_visual = RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado=RegistroVacaciones.ESTADO_APROBADA,
                    fecha_fin__year__gte=anio_saldo 
                ).aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0

                # Días acumulados restantes: lo que había menos lo consumido, piso 0
                dias_adicionales_iniciales = saldo.dias_adicionales or 0
                dias_acumulados = max(0, dias_adicionales_iniciales - consumo_visual)

                # Días restantes totales - columna "Restan"
                dias_restantes = saldo.dias_totales() - consumo_visual
                
                # Vacaciones en el rango TOTAL (aprobadas y pendientes)
                vacaciones = RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado__in=[RegistroVacaciones.ESTADO_APROBADA, RegistroVacaciones.ESTADO_PENDIENTE],
                    fecha_inicio__lte=fecha_fin_total,
                    fecha_fin__gte=fecha_inicio_total
                )
                
                # Obtener lista de vacaciones futuras aprobadas (para el selector PDF)
                vacaciones_futuras = list(RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado=RegistroVacaciones.ESTADO_APROBADA,
                    fecha_fin__gte=date.today()
                ).order_by('fecha_inicio'))
                
                empleados_list.append({
                    'empleado': emp,
                    'dias_disponibles': dias_disponibles,
                    'dias_acumulados': dias_acumulados,
                    'dias_restantes': dias_restantes,
                    'vacaciones': list(vacaciones),
                    'vacaciones_futuras': vacaciones_futuras,
                })
            
            if empleados_list:
                departamentos_data.append({
                    'departamento': depto,
                    'empleados': empleados_list
                })
        
        # Empleados sin departamento
        empleados_sin_depto = Empleado.objects.filter(
            departamento__isnull=True
        ).select_related('manager_aprobador').order_by('apellido', 'nombre')
        
        if empleados_sin_depto.exists():
            empleados_list = []
            for emp in empleados_sin_depto:
                saldo, created = SaldoVacaciones.objects.get_or_create(
                    empleado=emp,
                    ciclo=anio_saldo,
                    defaults={'dias_iniciales': emp.dias_base_lct(anio_saldo)}
                )
                
                # Días base del ciclo actual (sin acumulados) - columna "Disponible"
                dias_disponibles = saldo.dias_base_ciclo()
                
                # CORRECCIÓN: Calcular consumo visualmente para incluir vacaciones puente
                consumo_visual = RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado=RegistroVacaciones.ESTADO_APROBADA,
                    fecha_fin__year__gte=anio_saldo 
                ).aggregate(Sum('dias_solicitados'))['dias_solicitados__sum'] or 0

                # Días acumulados restantes: lo que había menos lo consumido, piso 0
                dias_adicionales_iniciales = saldo.dias_adicionales or 0
                dias_acumulados = max(0, dias_adicionales_iniciales - consumo_visual)

                # Días restantes totales - columna "Restan"
                dias_restantes = saldo.dias_totales() - consumo_visual
                
                vacaciones = RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado__in=[RegistroVacaciones.ESTADO_APROBADA, RegistroVacaciones.ESTADO_PENDIENTE],
                    fecha_inicio__lte=fecha_fin_total,
                    fecha_fin__gte=fecha_inicio_total
                )
                
                # Obtener lista de vacaciones futuras aprobadas (para el selector PDF)
                vacaciones_futuras = list(RegistroVacaciones.objects.filter(
                    empleado=emp,
                    estado=RegistroVacaciones.ESTADO_APROBADA,
                    fecha_fin__gte=date.today()
                ).order_by('fecha_inicio'))
                
                empleados_list.append({
                    'empleado': emp,
                    'dias_disponibles': dias_disponibles,
                    'dias_acumulados': dias_acumulados,
                    'dias_restantes': dias_restantes,
                    'vacaciones': list(vacaciones),
                    'vacaciones_futuras': vacaciones_futuras,
                })
            
            departamentos_data.append({
                'departamento': None,
                'empleados': empleados_list
            })

        context = {
            'anio_seleccionado': anio_param,
            'meses_data': meses_globales, # Lista unificada de meses
            'departamentos_data': departamentos_data,
        }

        return render(request, 'gestion/calendario_global.html', context)
        
    except Exception as e:
        import traceback
        return HttpResponse(f"<h1>Error: {e}</h1><pre>{traceback.format_exc()}</pre>")


@login_required
def exportar_calendario_excel(request):
    """
    Exporta el calendario de vacaciones a un archivo Excel.
    """
    try:
        anio_param = request.GET.get('anio', str(date.today().year))
        
        anios_a_mostrar = []
        if anio_param == 'todos':
            anios_a_mostrar = [2024, 2025, 2026]
        else:
            try:
                anios_a_mostrar = [int(anio_param)]
            except ValueError:
                anios_a_mostrar = [date.today().year]

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Calendario {anio_param}"

        # Estilos
        header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        week_fill = PatternFill(start_color="c2e9fb", end_color="c2e9fb", fill_type="solid")
        week_font = Font(bold=True, color="2c3e50", size=7)
        
        dept_fill = PatternFill(start_color="f5576c", end_color="f5576c", fill_type="solid")
        dept_font = Font(bold=True, color="FFFFFF", size=10)
        
        emp_fill = PatternFill(start_color="fcb69f", end_color="fcb69f", fill_type="solid")
        emp_font = Font(bold=False, color="2c3e50", size=10)
        
        disp_fill = PatternFill(start_color="fdcb6e", end_color="fdcb6e", fill_type="solid")
        acum_fill = PatternFill(start_color="56ab2f", end_color="56ab2f", fill_type="solid")
        rest_fill = PatternFill(start_color="9b59b6", end_color="9b59b6", fill_type="solid")
        
        vac_fill = PatternFill(start_color="38ef7d", end_color="38ef7d", fill_type="solid")
        vac_pending_fill = PatternFill(start_color="ffd93d", end_color="ffd93d", fill_type="solid")
        vac_font = Font(bold=True, color="000000")
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        # Bordes negros medianos para mejor visibilidad
        # Bordes finos negros (formato estándar de Excel)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Generar estructura de meses y semanas
        meses_globales = []
        for anio in anios_a_mostrar:
            datos_anio = _generar_datos_anio(anio)
            anio_corto = str(anio)[-2:]
            for mes in datos_anio['meses_data']:
                mes['nombre'] = f"{mes['nombre']} {anio_corto}"
                meses_globales.append(mes)

        # Calcular columnas
        col_offset = 5  # Empleado, Disponible, Acumuladas, Restan, (espacio)
        total_semanas = sum(len(mes['semanas']) for mes in meses_globales)

        # FILA 1: Headers de meses
        current_col = col_offset
        
        # Combinar verticalmente las columnas fijas (Empleado, Disponible, Acumuladas, Restan)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.cell(1, 1, "Empleado").fill = header_fill
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).alignment = center_align
        ws.cell(1, 1).border = thin_border
        
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        ws.cell(1, 2, "Disponible").fill = header_fill
        ws.cell(1, 2).font = header_font
        ws.cell(1, 2).alignment = center_align
        ws.cell(1, 2).border = thin_border
        
        ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        ws.cell(1, 3, "Acumuladas").fill = header_fill
        ws.cell(1, 3).font = header_font
        ws.cell(1, 3).alignment = center_align
        ws.cell(1, 3).border = thin_border
        
        ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        ws.cell(1, 4, "Restan").fill = header_fill
        ws.cell(1, 4).font = header_font
        ws.cell(1, 4).alignment = center_align
        ws.cell(1, 4).border = thin_border

        for mes in meses_globales:
            num_semanas = len(mes['semanas'])
            if num_semanas > 0:
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + num_semanas - 1)
                cell = ws.cell(1, current_col, mes['nombre'])
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                current_col += num_semanas

        # FILA 2: Headers de semanas
        ws.row_dimensions[2].height = 35  # Altura de la fila de semanas
        week_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_col = col_offset
        for mes in meses_globales:
            for semana in mes['semanas']:
                cell = ws.cell(2, current_col, semana['rango'])
                cell.fill = week_fill
                cell.font = week_font
                cell.alignment = week_align
                cell.border = thin_border
                current_col += 1

        # Obtener datos de empleados
        fecha_inicio_total = date(anios_a_mostrar[0], 1, 1)
        fecha_fin_total = date(anios_a_mostrar[-1], 12, 31)
        anio_saldo = date.today().year if date.today().year in anios_a_mostrar else anios_a_mostrar[0]

        departamentos = Departamento.objects.all().order_by('nombre')
        
        current_row = 3
        
        for depto in departamentos:
            empleados_depto = Empleado.objects.filter(
                departamento=depto
            ).select_related('manager_aprobador').order_by('apellido', 'nombre')
            
            if empleados_depto.exists():
                # Fila de departamento
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=col_offset + total_semanas - 1)
                cell = ws.cell(current_row, 1, depto.nombre.upper())
                cell.fill = dept_fill
                cell.font = dept_font
                cell.alignment = left_align
                current_row += 1
                
                # Empleados del departamento
                for emp in empleados_depto:
                    saldo, _ = SaldoVacaciones.objects.get_or_create(
                        empleado=emp,
                        ciclo=anio_saldo,
                        defaults={'dias_iniciales': emp.dias_base_lct(anio_saldo)}
                    )
                    
                    dias_disponibles = saldo.dias_base_ciclo()
                    dias_acumulados = saldo.dias_acumulados_restantes()
                    dias_restantes = saldo.total_disponible()
                    
                    vacaciones = RegistroVacaciones.objects.filter(
                        empleado=emp,
                        estado__in=[RegistroVacaciones.ESTADO_APROBADA, RegistroVacaciones.ESTADO_PENDIENTE],
                        fecha_inicio__lte=fecha_fin_total,
                        fecha_fin__gte=fecha_inicio_total
                    )
                    
                    # Columna Empleado
                    cell = ws.cell(current_row, 1, f"{emp.apellido}, {emp.nombre}")
                    cell.fill = emp_fill
                    cell.font = emp_font
                    cell.alignment = left_align
                    cell.border = thin_border
                    
                    # Columna Disponible
                    cell = ws.cell(current_row, 2, dias_disponibles)
                    cell.fill = disp_fill
                    cell.font = Font(bold=True, color="2c3e50")
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Columna Acumuladas
                    cell = ws.cell(current_row, 3, dias_acumulados)
                    cell.fill = acum_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Columna Restan
                    cell = ws.cell(current_row, 4, dias_restantes)
                    cell.fill = rest_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Marcar vacaciones en las semanas
                    current_col = col_offset
                    for mes in meses_globales:
                        for semana in mes['semanas']:
                            cell = ws.cell(current_row, current_col)
                            cell.border = thin_border
                            cell.alignment = center_align
                            
                            # Calcular días de vacaciones en esta semana y estado
                            dias_en_semana = 0
                            estado_vacacion = None
                            for dia in semana['dias']:
                                for vac in vacaciones:
                                    if vac.fecha_inicio <= dia <= vac.fecha_fin:
                                        dias_en_semana += 1
                                        # Priorizar estado APROBADO si hay mezcla
                                        if not estado_vacacion or estado_vacacion == RegistroVacaciones.ESTADO_PENDIENTE:
                                            estado_vacacion = vac.estado
                                        break
                            
                            if estado_vacacion == RegistroVacaciones.ESTADO_APROBADA:
                                cell.fill = vac_fill
                                cell.font = vac_font
                                cell.value = dias_en_semana
                            elif estado_vacacion == RegistroVacaciones.ESTADO_PENDIENTE:
                                cell.fill = vac_pending_fill
                                cell.font = vac_font
                                cell.value = dias_en_semana
                            
                            current_col += 1
                    
                    current_row += 1

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        
        for col in range(col_offset, col_offset + total_semanas):
            ws.column_dimensions[get_column_letter(col)].width = 5

        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"calendario_vacaciones_{anio_param}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        return HttpResponse(f"<h1>Error al exportar: {e}</h1><pre>{traceback.format_exc()}</pre>")


def _generar_datos_anio(anio_ciclo):
    """
    Función auxiliar para generar los datos del calendario de un año específico.
    """
    # 1. CALENDARIO ANUAL - Generar SEMANAS COMPLETAS del año (Lunes a Domingo)
    # Ajustamos para que la semana visual comience en LUNES.
    primer_dia = date(anio_ciclo, 1, 1)
    dia_semana = primer_dia.weekday()  # 0=Lunes, 6=Domingo
    
    # Calcular días hasta el Lunes anterior
    # Si es lunes(0) -> 0, Si es martes(1) -> 1 ...
    dias_hasta_lunes = dia_semana
    
    # Retroceder hasta el lunes
    fecha_inicio = primer_dia - timedelta(days=dias_hasta_lunes)
    
    # Generar todas las semanas del año
    todas_semanas = []
    fecha_actual = fecha_inicio
    
    # Continuar mientras la semana tenga al menos un día del año actual
    while True:
        # Crear semana de 7 días (Lunes a Domingo)
        semana = []
        for i in range(7):
            semana.append(fecha_actual + timedelta(days=i))
        
        # Verificar si esta semana tiene al menos un día del año actual
        tiene_dias_del_ano = any(dia.year == anio_ciclo for dia in semana)
        
        if not tiene_dias_del_ano and semana[0].year > anio_ciclo:
            break
        
        if tiene_dias_del_ano:
            # Determinar a qué mes asignar esta semana
            # Asignar al mes donde COMIENZA la semana (primer día = lunes)
            inicio = semana[0]
            fin = semana[6]
            
            # Si el inicio está en el año actual, usar ese mes
            # Si no, usar el mes del primer día del año actual en la semana
            if inicio.year == anio_ciclo:
                mes_principal = inicio.month
            else:
                # Buscar el primer día del año actual en la semana
                primer_dia_del_anio = next((dia for dia in semana if dia.year == anio_ciclo), None)
                if primer_dia_del_anio:
                    mes_principal = primer_dia_del_anio.month
                else:
                    continue  # No debería pasar, pero por seguridad
            
            # Formatear rango
            if inicio.month == fin.month and inicio.year == fin.year:
                rango = f"{inicio.day}-{fin.day}"
            elif inicio.year != fin.year:
                # Si cruza de año, mostrar año corto (ej: 29/12/24 - 4/1/25)
                rango = f"{inicio.day}/{inicio.month}/{str(inicio.year)[-2:]}-{fin.day}/{fin.month}/{str(fin.year)[-2:]}"
            else:
                rango = f"{inicio.day}/{inicio.month}-{fin.day}/{fin.month}"
            
            todas_semanas.append({
                'dias': semana,
                'inicio': inicio,
                'fin': fin,
                'mes': mes_principal,
                'rango': rango
            })
        
        fecha_actual += timedelta(days=7)
    
    # Agrupar semanas por mes
    meses_data = []
    for mes_num in range(1, 13):
        mes_nombre = MESES_ESPANOL.get(mes_num, f'Mes-{mes_num}')
        semanas_del_mes = [sem for sem in todas_semanas if sem['mes'] == mes_num]
        
        if semanas_del_mes:
            meses_data.append({
                'numero': mes_num,
                'nombre': mes_nombre,
                'semanas': semanas_del_mes,
                'total_semanas': len(semanas_del_mes)
            })
            
    return {
        'anio': anio_ciclo,
        'meses_data': meses_data,
    }




@login_required
@user_passes_test(is_manager)
def aprobar_rechazar_solicitud(request, solicitud_id):
    """
    Procesa la aprobación o el rechazo de una solicitud de vacaciones
    para un empleado y actualiza el saldo si se aprueba.
    """

    # Obtener la solicitud o devolver 404 si no existe
    solicitud = get_object_or_404(RegistroVacaciones, pk=solicitud_id)
    empleado = solicitud.empleado
    dias_solicitados = solicitud.dias_solicitados

    # Solo permitir la acción via POST
    if request.method != 'POST':
        messages.error(request, "Método no permitido. Utiliza el formulario.")
        return redirect('gestion:historial_global')

    accion = request.POST.get('accion')  # Debe ser 'aprobar' o 'rechazar'

    try:
        # Obtener el Empleado asociado al usuario actual (manager)
        try:
            manager_empleado = Empleado.objects.get(user=request.user)
        except Empleado.DoesNotExist:
            messages.error(request, "Error: Tu usuario no está asociado a un perfil de empleado.")
            logger.error(f"Usuario {request.user.username} no tiene perfil de Empleado asociado")
            return redirect('gestion:historial_global')

        # Usar transacciones atómicas para asegurar que el saldo y la solicitud se actualicen juntos
        with transaction.atomic():
            # 1. Validaciones específicas según la acción
            if accion in ['aprobar', 'rechazar'] and solicitud.estado != RegistroVacaciones.ESTADO_PENDIENTE:
                 messages.warning(request, f"La solicitud ya fue procesada y está en estado: '{solicitud.estado}'.")
                 return redirect('gestion:historial_global')

            # Para cancelar, permitimos si está Aprobada (o incluso Pendiente si se desea)
            if accion == 'cancelar' and solicitud.estado not in [RegistroVacaciones.ESTADO_APROBADA, RegistroVacaciones.ESTADO_PENDIENTE]:
                 messages.warning(request, f"No se puede cancelar una solicitud en estado: '{solicitud.estado}'.")
                 return redirect('gestion:historial_global')

            # 2. Obtener o crear el saldo de vacaciones para el ciclo (año) actual
            ciclo_actual = datetime.now().year
            saldo, created = SaldoVacaciones.objects.get_or_create(
                empleado=empleado,
                ciclo=ciclo_actual,
                defaults={'dias_iniciales': empleado.dias_base_lct(ciclo_actual)}
            )

            if accion == 'aprobar':
                # 3. Acción de aprobar: Validar saldo
                if saldo.total_disponible() < dias_solicitados:
                    messages.error(
                        request,
                        f"Saldo insuficiente. {empleado.nombre} tiene {saldo.total_disponible()} días disponibles y solicita {dias_solicitados}."
                    )
                    raise Exception("Fallo en la aprobación: Saldo insuficiente.")

                # NOTA: No es necesario descontar manualmente los días porque el modelo
                # SaldoVacaciones calcula automáticamente los días consumidos.
                
                # Actualizar estado de la solicitud y guardar el objeto RegistroVacaciones
                solicitud.estado = RegistroVacaciones.ESTADO_APROBADA
                solicitud.manager_aprobador = manager_empleado
                solicitud.fecha_aprobacion = date.today()
                solicitud.save()
                
                # Enviar notificación por email al empleado
                enviar_email_cambio_estado(request, solicitud)

                # Notificación interna
                crear_notificacion(
                    usuario=solicitud.empleado.user,
                    titulo="Vacaciones Aprobadas ✅",
                    mensaje=f"Tu solicitud para el ciclo {datetime.now().year} ha sido APROBADA.",
                    url="gestion:historial_personal",
                    solicitud=solicitud
                )
                
                # Auto-limpiar notificaciones PENDIENTES de esta solicitud para el manager
                Notificacion.objects.filter(solicitud=solicitud, usuario=request.user).update(leida=True)

                messages.success(
                    request,
                    f"Vacaciones de {empleado.nombre} APROBADAS. Se descontaron {dias_solicitados} días. Nuevo saldo: {saldo.total_disponible()} días."
                )

            elif accion == 'rechazar':
                # 4. Acción de rechazar: Actualizar estado sin modificar saldo
                solicitud.estado = RegistroVacaciones.ESTADO_RECHAZADA
                solicitud.manager_aprobador = manager_empleado
                solicitud.fecha_aprobacion = date.today()
                solicitud.save()

                # Enviar notificación por email al empleado
                enviar_email_cambio_estado(request, solicitud)

                # Notificación interna
                crear_notificacion(
                    usuario=solicitud.empleado.user,
                    titulo="Solicitud Rechazada ❌",
                    mensaje=f"Tu solicitud para el ciclo {datetime.now().year} ha sido RECHAZADA.",
                    url="gestion:historial_personal",
                    solicitud=solicitud
                )

                # Auto-limpiar notificaciones PENDIENTES de esta solicitud para el manager
                Notificacion.objects.filter(solicitud=solicitud, usuario=request.user).update(leida=True)

                messages.warning(request, f"Vacaciones de {empleado.nombre} RECHAZADAS. El saldo no fue afectado.")

            elif accion == 'cancelar':
                # NUEVA ACCIÓN: Cancelar solicitud (incluso si estaba aprobada)
                # Al cambiar el estado a Cancelada, el cálculo dinámico de saldo
                # automáticamente dejará de contar estos días como consumidos.
                estado_anterior = solicitud.estado
                solicitud.estado = RegistroVacaciones.ESTADO_CANCELADA
                # Mantenemos o actualizamos el manager que canceló
                solicitud.manager_aprobador = manager_empleado 
                solicitud.save()
                
                # Enviar notificación por email al empleado
                enviar_email_cambio_estado(request, solicitud)

                # Notificación interna
                crear_notificacion(
                    usuario=solicitud.empleado.user,
                    titulo="Solicitud Cancelada ⚠️",
                    mensaje=f"Tu solicitud para el ciclo {datetime.now().year} ha sido CANCELADA y los días devueltos.",
                    url="gestion:historial_personal",
                    solicitud=solicitud
                )

                # Auto-limpiar notificaciones PENDIENTES
                Notificacion.objects.filter(solicitud=solicitud).update(leida=True)

                msg_extra = ""
                if estado_anterior == RegistroVacaciones.ESTADO_APROBADA:
                    msg_extra = " Los días descontados han sido devueltos al saldo."
                
                messages.success(request, f"Solicitud de {empleado.nombre} CANCELADA.{msg_extra}")

            else:
                # 5. Acción inválida
                messages.error(request, f"Acción '{accion}' inválida o no reconocida.")
                raise Exception("Acción de formulario no válida.")

    except Exception as e:
        logger.error(f"Error procesando solicitud {solicitud_id}: {e}")
        if not str(e).startswith("Fallo en la aprobación: Saldo insuficiente."):
             messages.error(request, "Error interno al procesar la solicitud. Contacta a soporte.")
    
    return redirect('gestion:historial_global')


@login_required
def solicitar_vacaciones(request):
    """
    Vista HÍBRIDA (Auto-Servicio + Manager):
    - Si es Manager: Puede seleccionar cualquier empleado para cargarle vacaciones.
    - Si es Empleado: Solo puede cargarse a sí mismo (Auto-Servicio).
    """
    try:
        # Obtener el perfil del usuario actual
        mi_empleado = Empleado.objects.get(user=request.user)
        es_manager = mi_empleado.es_manager
    except Empleado.DoesNotExist:
        messages.error(request, "Tu usuario no tiene un perfil de empleado asociado.")
        return redirect('dashboard')

    current_year = timezone.now().year
    
    # --- Lógica de Selección de Empleado ---
    # Si es manager, tomamos el ID del POST/GET o None.
    # Si es empleado, forzamos su propio ID.
    empleado_seleccionado_id = None
    
    if es_manager:
        empleado_seleccionado_id = request.POST.get('empleado_id', request.GET.get('empleado_id'))
        context_empleados = Empleado.objects.all().order_by('apellido', 'nombre')
    else:
        empleado_seleccionado_id = mi_empleado.id
        context_empleados = [mi_empleado] # Solo él mismo en la lista
    
    # -----------------------------------------------------------
    #   CONTEXTO BASE
    # -----------------------------------------------------------
    current_year = timezone.now().year
    start_goce = date(current_year, 10, 1)
    end_goce = date(current_year + 1, 4, 30)
    
    context = {
        'es_manager': es_manager, # Para mostrar/ocultar selector en HTML
        'mi_empleado': mi_empleado,
        'hoy': date.today().isoformat(),
        'todos_empleados': context_empleados,
        'empleado_seleccionado_id': str(empleado_seleccionado_id) if empleado_seleccionado_id else "",
        'periodo_goce_inicio': start_goce.strftime("%d/%m/%Y"),
        'periodo_goce_fin': end_goce.strftime("%d/%m/%Y"),
        'estado_a_registrar': 'Pendiente',
        'saldo_disponible': 'N/A'
    }

    # -----------------------------------------------------------
    #   CARGAR SALDO Y KPI (SI HAY EMPLEADO SELECCIONADO)
    # -----------------------------------------------------------
    if empleado_seleccionado_id:
        try:
            empleado_afectado = Empleado.objects.get(pk=empleado_seleccionado_id)
            
            # Seguridad: Si NO es manager, verificar que no esté intentando ver datos de otro
            if not es_manager and empleado_afectado != mi_empleado:
                messages.error(request, "Acción no autorizada. Solo puedes ver tus propios datos.")
                return redirect('gestion:solicitar_vacaciones')

            saldo, created = SaldoVacaciones.objects.get_or_create(
                empleado=empleado_afectado,
                ciclo=current_year,
                defaults={'dias_iniciales': empleado_afectado.dias_base_lct(current_year)}
            )

            context['saldo_disponible'] = f"{saldo.total_disponible()} días"
            
            # --- KPIs para las Cards ---
            dias_totales = (saldo.dias_iniciales or 0) + (saldo.dias_adicionales or 0)
            
            dias_usados = RegistroVacaciones.objects.filter(
                empleado=empleado_afectado,
                estado=RegistroVacaciones.ESTADO_APROBADA,
                fecha_inicio__year=current_year
            ).aggregate(total=Sum('dias_solicitados'))['total'] or 0
            
            dias_pendientes = RegistroVacaciones.objects.filter(
                empleado=empleado_afectado,
                estado=RegistroVacaciones.ESTADO_PENDIENTE,
                fecha_inicio__year=current_year
            ).aggregate(total=Sum('dias_solicitados'))['total'] or 0

            context.update({
                'dias_totales': dias_totales,
                'dias_usados': dias_usados,
                'dias_pendientes': dias_pendientes,
                'dias_acumulados': saldo.dias_acumulados_restantes()
            })

            if created:
                if es_manager:
                    messages.warning(request, f"Se inicializó el saldo de {empleado_afectado.nombre} automáticamente.")

        except Empleado.DoesNotExist:
            pass

    # -----------------------------------------------------------
    #   PROCESAR FORMULARIO POST
    # -----------------------------------------------------------
    if request.method == 'POST':
        data = request.POST
        
        # 1. Determinar Empleado Afectado (Robustez)
        target_id = data.get('empleado_id')
        
        if es_manager:
            if not target_id:
                messages.error(request, "Debe seleccionar un empleado.")
                return render(request, 'gestion/solicitud.html', context)
            empleado_afectado = get_object_or_404(Empleado, pk=target_id)
        else:
            # Si es empleado, ignoramos lo que venga en el POST y usamos al usuario actual
            empleado_afectado = mi_empleado
            target_id = mi_empleado.id
        
        # 2. Validar Fechas
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return render(request, 'gestion/solicitud.html', context)

        if fecha_inicio > fecha_fin:
            messages.error(request, "La fecha de inicio no puede ser mayor que la fecha fin.")
            return render(request, 'gestion/solicitud.html', context)

        # 3. Validar Saldo
        saldo, created = SaldoVacaciones.objects.get_or_create(
            empleado=empleado_afectado,
            ciclo=current_year,
            defaults={'dias_iniciales': empleado_afectado.dias_base_lct(current_year)}
        )

        saldo_disponible = saldo.total_disponible()
        dias_solicitados = (fecha_fin - fecha_inicio).days + 1

        if dias_solicitados > saldo_disponible:
            messages.error(
                request,
                f"Saldo insuficiente. Solicitas {dias_solicitados} días y tienes {saldo_disponible} disponibles."
            )
            return render(request, 'gestion/solicitud.html', context)

        # 4. Guardar Solicitud
        try:
            with transaction.atomic():
                solicitud = RegistroVacaciones.objects.create(
                    empleado=empleado_afectado,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    estado=RegistroVacaciones.ESTADO_PENDIENTE,
                    manager_aprobador=None, # Siempre null al inicio (pendiente)
                    razon=data.get('razon')
                )
            
            # --- NOTIFICACIÓN REAL ---
            enviar_email_nueva_solicitud(request, solicitud)
            
            messages.success(
                request,
                f"Solicitud enviada correctamente ({dias_solicitados} días). Esperando aprobación."
            )
            
            # Redirección inteligente
            if es_manager:
                return redirect('gestion:historial_global')
            else:
                return redirect('gestion:historial_personal') # O dashboard

        except Exception as e:
            messages.error(request, f"Error al guardar solicitud: {e}")
            return render(request, 'gestion/solicitud.html', context)

    # ============================================================
    #   GET → Render inicial
    # ============================================================
    return render(request, 'gestion/solicitud.html', context)


@login_required
@user_passes_test(is_manager)
def exportar_notificacion_vacaciones_pdf(request, empleado_id, vacacion_id):
    """
    Genera un PDF PREMIUM de notificación de vacaciones:
    - Logo institucional
    - Encabezado corporativo
    - Datos del empleado
    - Fechas de vacaciones
    - Firma del responsable
    """

    # Obtener los datos
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    registro = get_object_or_404(RegistroVacaciones, pk=vacacion_id)

    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"Notificacion_Vacaciones_{empleado.apellido}_{empleado.nombre}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    """
    Genera un PDF con formato de FORMULARIO TRADICIONAL (Legacy):
    - Título centrado
    - Campos detallados (Legajo, Sector, Periodo, Restan)
    - Firmas lado a lado
    - Checkbox de anticipo
    """

    # Obtener los datos
    empleado = get_object_or_404(Empleado, pk=empleado_id)
    registro = get_object_or_404(RegistroVacaciones, pk=vacacion_id)

    # Calcular saldo restante
    # FIX: Filtrar por empleado Y por el año (ciclo) de la vacación para evitar "MultipleObjectsReturned"
    anio_ciclo = registro.fecha_inicio.year
    saldo_qs = SaldoVacaciones.objects.filter(empleado=empleado, ciclo=anio_ciclo)
    
    # Si existe saldo para ese año, lo usamos. Si no, tomamos el primero que aparezca o 0.
    if saldo_qs.exists():
        saldo = saldo_qs.first()
        dias_restantes = saldo.total_disponible()
    else:
        # Fallback: Intentar obtener el saldo del año actual si no coincide el de la vacación
        saldo_actual = SaldoVacaciones.objects.filter(empleado=empleado, ciclo=datetime.now().year).first()
        dias_restantes = saldo_actual.total_disponible() if saldo_actual else 0

    # Calcular fecha de retoma (día siguiente al fin)
    fecha_retoma = registro.fecha_fin + timedelta(days=1)

    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"Notificacion_Vacaciones_{empleado.apellido}_{empleado.nombre}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Configuración de página
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    top_y = height - 50  # Margen superior inicial

    # --- TÍTULO ---
    pdf.setFont("Helvetica-Bold", 16)
    title = "NOTIFICACION DE VACACIONES"
    title_width = pdf.stringWidth(title, "Helvetica-Bold", 16)
    pdf.drawString((width - title_width) / 2, top_y, title)
    
    # Línea debajo del título
    pdf.setLineWidth(1)
    pdf.line(50, top_y - 10, width - 50, top_y - 10)

    # --- CAMPOS DEL FORMULARIO ---
    current_y = top_y - 50
    left_margin = 55
    value_x = 200 # X coordinate alignment for some values

    # FECHA
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "FECHA:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(130, current_y, datetime.today().strftime('%d/%m/%Y'))
    
    current_y -= 30

    # APELLIDO Y NOMBRE
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "APELLIDO Y NOMBRE:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(220, current_y, f"{empleado.apellido.upper()} {empleado.nombre.upper()}")

    current_y -= 30

    # LEGAJO
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "LEGAJO:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(130, current_y, str(empleado.legajo or ""))

    current_y -= 30

    # SECTOR
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "SECTOR:")
    pdf.setFont("Helvetica", 11)
    sector = empleado.departamento.nombre.upper() if empleado.departamento else ""
    pdf.drawString(130, current_y, sector)

    current_y -= 30

    # PERIODO DE VACACIONES (AÑO)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "PERIODO DE VACACIONES (AÑO):")
    pdf.setFont("Helvetica", 11)
    # Asumimos que el periodo es el año de inicio de la vacación o el actual
    anio_periodo = registro.fecha_inicio.year 
    pdf.drawString(280, current_y, str(anio_periodo))

    current_y -= 30

    # DIAS A TOMAR
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "DIAS A TOMAR (CORRIDOS):")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(280, current_y, str(registro.dias_solicitados))

    current_y -= 30

    # RESTAN DIAS
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "RESTAN DIAS DEL PERIODO") # Label exact match to image
    pdf.setFont("Helvetica", 11)
    pdf.drawString(280, current_y, str(dias_restantes))

    current_y -= 40 # Extra space before Dates

    # DESDE / HASTA (Misma línea)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "DESDE:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(110, current_y, registro.fecha_inicio.strftime('%d/%m/%Y'))

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(260, current_y, "HASTA INCLUSIVE:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(390, current_y, registro.fecha_fin.strftime('%d/%m/%Y'))

    current_y -= 40

    # RETOMA
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "RETOMA A SUS TAREAS EL DIA:")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(280, current_y, fecha_retoma.strftime('%d/%m/%Y'))

    current_y -= 40

    # ANTICIPO DE SUELDO
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(left_margin, current_y, "ANTICIPO DE SUELDO:")
    
    # Checkboxes ficticios (cuadraditos)
    pdf.setFont("Helvetica", 14) 
    pdf.rect(230, current_y, 12, 12, fill=0) # Box SI
    pdf.setFont("Helvetica", 11)
    pdf.drawString(250, current_y + 2, "SI")

    pdf.rect(300, current_y, 12, 12, fill=0) # Box NO
    pdf.drawString(320, current_y + 2, "NO")

    current_y -= 25
    pdf.setFont("Helvetica", 7)
    pdf.drawString(left_margin, current_y, "POR FAVOR MARCAR CON UNA X EL QUE NO CORRESPONDA.")

    # --- FIRMAS ---
    current_y -= 100 # Espacio para firmas

    # Columna Izquierda: Interesado
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left_margin, current_y, "FIRMA DEL INTERESADO:")
    pdf.line(left_margin, current_y - 40, left_margin + 200, current_y - 40)
    
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left_margin, current_y - 60, "ACLARACION:")
    pdf.line(left_margin, current_y - 80, left_margin + 200, current_y - 80)

    # Columna Derecha: Supervisor
    right_col_x = 320
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(right_col_x, current_y, "FIRMA DEL SUPERVISOR:")
    pdf.line(right_col_x, current_y - 40, right_col_x + 200, current_y - 40)

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(right_col_x, current_y - 60, "ACLARACION:")
    pdf.line(right_col_x, current_y - 80, right_col_x + 200, current_y - 80)

    # --- COMENTARIOS ---
    current_y -= 130
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left_margin, current_y, "COMENTARIOS")
    pdf.line(left_margin, current_y - 5, width - 50, current_y - 5)

    # Finalizar PDF
    pdf.showPage()
    pdf.save()

    return response


@login_required
def solicitar_mis_vacaciones(request):
    """
    Vista simplificada para que el EMPLEADO solicite sus propias vacaciones.
    Sin selección de usuario (es él mismo).
    """
    try:
        if not hasattr(request.user, 'empleado'):
             messages.error(request, "No tienes un perfil de empleado asociado.")
             return redirect('gestion:dashboard')
        empleado = request.user.empleado
    except Empleado.DoesNotExist:
        messages.error(request, "No tienes un perfil de empleado asociado.")
        return redirect('gestion:dashboard')

    current_year = timezone.now().year
    
    # Obtener saldo
    saldo, created = SaldoVacaciones.objects.get_or_create(
        empleado=empleado,
        ciclo=current_year,
        defaults={'dias_iniciales': empleado.dias_base_lct(current_year)}
    )
    saldo_disponible = saldo.total_disponible()

    if request.method == 'POST':
        fecha_inicio_str = request.POST.get('fecha_inicio')
        fecha_fin_str = request.POST.get('fecha_fin')

        try:
            fecha_inicio = date.fromisoformat(fecha_inicio_str)
            fecha_fin = date.fromisoformat(fecha_fin_str)
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('gestion:solicitar_mis_vacaciones')

        if fecha_inicio > fecha_fin:
            messages.error(request, "La fecha de inicio no puede ser mayor que la fecha fin.")
            return redirect('gestion:solicitar_mis_vacaciones')

        dias_solicitados = (fecha_fin - fecha_inicio).days + 1

        if dias_solicitados > saldo_disponible:
            messages.error(
                request, 
                f"Saldo insuficiente. Solicitas {dias_solicitados} días y tienes {saldo_disponible}."
            )
            return redirect('gestion:solicitar_mis_vacaciones')

        # Crear solicitud PENDIENTE
        try:
            with transaction.atomic():
                solicitud = RegistroVacaciones.objects.create(
                    empleado=empleado,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    estado=RegistroVacaciones.ESTADO_PENDIENTE,
                    manager_aprobador=empleado.manager_aprobador, # Se asigna a su manager
                    razon="Solicitud personal desde App"
                )
            
            # Enviar notificación por email
            enviar_email_nueva_solicitud(request, solicitud)

            # Notificación interna para el manager y para todos los Administradores
            notificados = set()
            if solicitud.manager_aprobador and solicitud.manager_aprobador.user:
                crear_notificacion(
                    usuario=solicitud.manager_aprobador.user,
                    titulo="Nueva Solicitud de Vacaciones 🔔",
                    mensaje=f"{empleado.nombre} {empleado.apellido} ha solicitado vacaciones.",
                    url="gestion:aprobacion_manager",
                    solicitud=solicitud
                )
                notificados.add(solicitud.manager_aprobador.user.id)
            
            # Notificar también a todos los Administradores (superusuarios) que no sean el manager ya notificado
            admin_users = User.objects.filter(is_superuser=True)
            for a_user in admin_users:
                if a_user.id not in notificados:
                    crear_notificacion(
                        usuario=a_user,
                        titulo="Seguimiento: Nueva Solicitud 🔔",
                        mensaje=f"{empleado.nombre} {empleado.apellido} ha solicitado vacaciones.",
                        url="gestion:aprobacion_manager",
                        solicitud=solicitud
                    )

            
            messages.success(request, "¡Solicitud enviada con éxito! Tu manager deberá aprobarla.")

            return redirect('gestion:dashboard')
        except Exception as e:
            messages.error(request, f"Error al crear solicitud: {str(e)}")
            return redirect('gestion:solicitar_mis_vacaciones')

    context = {
        'saldo_disponible': saldo_disponible,
        'current_year': current_year,
        'hoy': date.today().isoformat()
    }
    return render(request, 'gestion/solicitar_personal.html', context)


from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

@login_required
def cambiar_password(request):
    """
    Vista para obligar al usuario a cambiar su contraseña (Primer Login).
    """
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Mantener la sesión activa después del cambio de password
            update_session_auth_hash(request, user)  
            
            # Actualizar flag en Empleado
            try:
                if hasattr(user, 'empleado'):
                    empleado = user.empleado
                    empleado.primer_login = False
                    
                    # Manejo seguro si el modelo cambió recientemente y la instancia está 'stale'
                    # Aunque en este contexto request.user.empleado debería estar fresco o ser refrescable
                    empleado.save(update_fields=['primer_login'])
            except Exception as e:
                # Si falla guardar el flag, loguear pero permitir continuar
                logger.error(f"Error al actualizar primer_login para {user.username}: {e}")
                pass
                
            messages.success(request, '¡Contraseña actualizada con éxito! Bienvenido.')
            return redirect('gestion:dashboard')
        else:
            messages.error(request, 'Por favor corrija los errores indicados.')
    else:
        form = PasswordChangeForm(request.user)
        
    return render(request, 'gestion/cambiar_password.html', {
        'form': form
    })

@login_required
def lista_notificaciones(request):
    """Muestra todas las notificaciones del usuario y sus tareas pendientes."""
    notificaciones = request.user.notificaciones.all().order_by('-fecha_creacion')
    
    # Si es manager, traer también solicitudes pendientes para mostrar en el mismo listado
    solicitudes_pendientes = []
    if hasattr(request.user, 'empleado') and request.user.empleado.es_manager:
        from django.db.models import Q
        # Si es superusuario (administrador), ve todas. Si no, solo las suyas.
        if request.user.is_superuser:
            filter_q = Q()
        else:
            filter_q = Q(manager_aprobador=request.user.empleado)
            
        solicitudes_pendientes = RegistroVacaciones.objects.filter(
            filter_q,
            estado=RegistroVacaciones.ESTADO_PENDIENTE
        ).order_by('fecha_solicitud')

    # Si viene por POST, marcar todas como leídas
    if request.method == 'POST' and 'marcar_todas' in request.POST:
        request.user.notificaciones.filter(leida=False).update(leida=True)
        messages.success(request, "Todas las notificaciones marcadas como leídas.")
        return redirect('gestion:lista_notificaciones')
        
    context = {
        'notificaciones': notificaciones,
        'solicitudes_pendientes': solicitudes_pendientes
    }
    return render(request, 'gestion/notificaciones.html', context)


@login_required
def marcar_notificacion_leida(request, notif_id):
    """Marca una notificación como leída y redirige a su URL."""
    notificacion = get_object_or_404(Notificacion, id=notif_id, usuario=request.user)
    
    # LÓGICA INTELIGENTE: Si la notificación tiene una solicitud vinculada y es una tarea pendiente,
    # NO la marcamos como leída automáticamente solo por hacer clic, a menos que el usuario lo haga manual.
    # Pero para no marear, vamos a marcarla como leída solo si NO es de tipo "Nueva Solicitud"
    # o si el usuario explícitamente quiere ir a la URL.
    
    # Por ahora: Si es de tipo Solicitud Pendiente, el manager querrá mantenerla hasta resolverla.
    is_task = notificacion.solicitud and notificacion.solicitud.estado == RegistroVacaciones.ESTADO_PENDIENTE
    
    if not is_task:
        notificacion.leida = True
        notificacion.save()
    
    if notificacion.url:

        try:
            # Si el URL contiene el nombre de una vista de Django, intentar resolverla
            from django.urls import reverse
            return redirect(reverse(notificacion.url))
        except:
            # Si es un path directo, redirigir
            return redirect(notificacion.url)
    
    return redirect('gestion:lista_notificaciones')

@login_required
def api_check_notificaciones(request):
    """Endpoint para polling de notificaciones en tiempo real."""
    last_notif_id = request.GET.get('last_id', 0)
    
    try:
        last_notif_id = int(last_notif_id)
    except ValueError:
        last_notif_id = 0

    # 1. Notificaciones nuevas (creadas después del last_id)
    nuevas_notif = request.user.notificaciones.filter(id__gt=last_notif_id).order_by('id')
    
    # 2. Conteo total de no leídas
    unread_count = request.user.notificaciones.filter(leida=False).count()
    
    # 3. Tareas pendientes (solo para managers)
    tareas_pendientes = 0
    if hasattr(request.user, 'empleado') and request.user.empleado.es_manager:
        from django.db.models import Q
        # Si es Administrador (Superuser), ve TODAS las pendientes
        if request.user.is_superuser:
            filter_q = Q()
        else:
            filter_q = Q(manager_aprobador=request.user.empleado)
            
        tareas_pendientes = RegistroVacaciones.objects.filter(
            filter_q,
            estado=RegistroVacaciones.ESTADO_PENDIENTE
        ).count()

    notif_data = []
    max_id = last_notif_id
    for n in nuevas_notif:
        notif_data.append({
            'id': n.id,
            'titulo': n.titulo,
            'mensaje': n.mensaje,
            'url': n.url
        })
        max_id = max(max_id, n.id)

    return JsonResponse({
        'unread_count': unread_count + tareas_pendientes,
        'tareas_pendientes': tareas_pendientes,
        'nuevas': notif_data,
        'last_id': max_id
    })


